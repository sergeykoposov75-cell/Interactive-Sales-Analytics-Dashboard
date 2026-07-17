# -*- coding: utf-8 -*-
"""
Генератор дашборда продаж для розничной компании.
Создаёт Excel-файл с демо-данными, сводками, диаграммами и интерактивными фильтрами.
"""

import numpy as np
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os
import sys

# ────────────────────────── КОНСТАНТЫ ──────────────────────────
SEED = 42
np.random.seed(SEED)

REGIONS = ["Север", "Юг", "Центр", "Восток", "Запад"]
CATEGORIES = ["Электроника", "Одежда", "Продукты", "Мебель", "Косметика"]
REGION_WEIGHTS = np.array([0.18, 0.20, 0.25, 0.15, 0.22])
CAT_PROBS = np.array([0.25, 0.20, 0.25, 0.10, 0.20])

CAT_PARAMS = [
    (5000, 50000, 1, 10),   # Электроника
    (2000, 30000, 2, 20),   # Одежда
    (1000, 15000, 5, 50),   # Продукты
    (10000, 50000, 1, 5),   # Мебель
    (1000, 25000, 1, 15),   # Косметика
]

MONTH_LABELS = ["Янв", "Фев", "Мар", "Апр", "Май", "Июн",
                "Июл", "Авг", "Сен", "Окт", "Ноя", "Дек"]

BLUE = "1E88E5"
ORANGE = "FFB74D"
DARK_BLUE = "0D47A1"
LIGHT_BLUE = "E3F2FD"
BORDER_COLOR = "BDBDBD"

BLUE_FILL = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
ORANGE_FILL = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

TITLE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
SECTION_FONT = Font(name="Calibri", bold=True, color=DARK_BLUE, size=12)
KPI_LABEL_FONT = Font(name="Calibri", bold=True, color=DARK_BLUE, size=9)
KPI_VALUE_FONT = Font(name="Calibri", bold=True, color="333333", size=14)
BODY_FONT = Font(name="Calibri", color="333333", size=10)
BODY_BOLD = Font(name="Calibri", bold=True, color="333333", size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_CENTER = Alignment(horizontal="left", vertical="center")


# ────────────────────────── ГЕНЕРАЦИЯ ДАННЫХ ──────────────────────────
def generate_data():
    """Генерирует синтетические данные о продажах за 2025 год.
    Возвращает список словарей."""
    start = date(2025, 1, 1)
    end = date(2025, 12, 31)
    records = []
    d = start
    while d <= end:
        wd = d.weekday()
        day_factor = 1.2 if wd >= 5 else 1.0
        if d.day >= 25:
            day_factor *= 1.15
        if d.month == 12 and d.day >= 15:
            day_factor *= 1.4
        elif d.month == 1 and d.day <= 10:
            day_factor *= 1.2

        n = max(5, min(45, int(np.random.randint(12, 30) * day_factor)))

        for _ in range(n):
            reg_idx = int(np.random.choice(len(REGIONS), p=REGION_WEIGHTS))
            cat_idx = int(np.random.choice(len(CATEGORIES), p=CAT_PROBS))
            min_r, max_r, min_t, max_t = CAT_PARAMS[cat_idx]
            rev = int(np.random.randint(min_r, max_r + 1))
            trx = int(np.random.randint(min_t, max_t + 1))
            avg = round(rev / trx, 2)
            records.append({
                "date": d,
                "region": REGIONS[reg_idx],
                "category": CATEGORIES[cat_idx],
                "revenue": rev,
                "transactions": trx,
                "avg_check": avg,
            })
        d += timedelta(days=1)
    return records


def compute_summaries(records):
    """Считает агрегированные сводки из списка записей."""
    rev_tot = sum(r["revenue"] for r in records)
    trx_tot = sum(r["transactions"] for r in records)

    # По регионам
    reg_data = {reg: {"revenue": 0, "transactions": 0} for reg in REGIONS}
    for r in records:
        reg_data[r["region"]]["revenue"] += r["revenue"]
        reg_data[r["region"]]["transactions"] += r["transactions"]
    reg_summary = []
    for reg in REGIONS:
        rev = reg_data[reg]["revenue"]
        trx = reg_data[reg]["transactions"]
        avg = round(rev / trx, 2) if trx else 0
        reg_summary.append((reg, rev, trx, avg))
    reg_summary.sort(key=lambda x: x[1], reverse=True)

    # По категориям
    cat_data = {cat: {"revenue": 0, "transactions": 0} for cat in CATEGORIES}
    for r in records:
        cat_data[r["category"]]["revenue"] += r["revenue"]
        cat_data[r["category"]]["transactions"] += r["transactions"]
    cat_summary = []
    for cat in CATEGORIES:
        rev = cat_data[cat]["revenue"]
        trx = cat_data[cat]["transactions"]
        avg = round(rev / trx, 2) if trx else 0
        cat_summary.append((cat, rev, trx, avg))
    cat_summary.sort(key=lambda x: x[1], reverse=True)

    # По месяцам
    monthly = {m: {"revenue": 0, "transactions": 0} for m in range(1, 13)}
    for r in records:
        m = r["date"].month
        monthly[m]["revenue"] += r["revenue"]
        monthly[m]["transactions"] += r["transactions"]
    monthly_list = []
    for m in range(1, 13):
        monthly_list.append((
            MONTH_LABELS[m - 1],
            monthly[m]["revenue"],
            monthly[m]["transactions"],
        ))

    return reg_summary, cat_summary, monthly_list, rev_tot, trx_tot


# ────────────────────────── ФОРМАТИРОВАНИЕ ──────────────────────────
def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = BLUE_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def write_data_rows(ws, data, start_row, fmt_map=None):
    """Записывает список кортежей/списков начиная со start_row и форматирует."""
    for i, row_data in enumerate(data):
        r = start_row + i
        for j, val in enumerate(row_data):
            cell = ws.cell(row=r, column=j + 1, value=val)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            if fmt_map and (j + 1) in fmt_map:
                cell.number_format = fmt_map[j + 1]


# ────────────────────────── СОЗДАНИЕ EXCEL ──────────────────────────
def create_workbook(records, reg_summary, cat_summary, monthly_list,
                    total_revenue, total_transactions):
    wb = Workbook()

    # ════════════════ ЛИСТ 1: ДАННЫЕ ════════════════
    ws = wb.active
    ws.title = "Данные"
    headers = ["Дата", "Регион", "Категория", "Выручка", "Транзакции", "Средний чек"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, 6)

    for i, r in enumerate(records):
        row = i + 2
        ws.cell(row=row, column=1, value=r["date"]).number_format = "DD.MM.YYYY"
        ws.cell(row=row, column=2, value=r["region"])
        ws.cell(row=row, column=3, value=r["category"])
        ws.cell(row=row, column=4, value=r["revenue"]).number_format = '#,##0'
        ws.cell(row=row, column=5, value=r["transactions"]).number_format = '#,##0'
        ws.cell(row=row, column=6, value=r["avg_check"]).number_format = '#,##0.00'
        for c in range(1, 7):
            cell = ws.cell(row=row, column=c)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER

    last_row = len(records) + 1
    ws.auto_filter.ref = f"A1:F{last_row}"
    ws.freeze_panes = "A2"
    for i, w in enumerate([14, 12, 16, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ════════════════ ЛИСТ 2: СВОДКА ПО РЕГИОНАМ ════════════════
    ws2 = wb.create_sheet("Сводка по регионам")
    h2 = ["Регион", "Выручка", "Транзакции", "Средний чек"]
    for c, h in enumerate(h2, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, 4)
    write_data_rows(ws2, reg_summary, 2, {2: '#,##0', 3: '#,##0', 4: '#,##0.00'})
    lr2 = len(reg_summary) + 1
    ws2.auto_filter.ref = f"A1:D{lr2}"
    ws2.freeze_panes = "A2"
    ws2.conditional_formatting.add(
        f"A2:A{lr2}",
        FormulaRule(formula=[f"B2=MAX($B$2:$B${lr2})"],
                     fill=ORANGE_FILL, font=Font(bold=True, color="FFFFFF")),
    )
    for i, w in enumerate([12, 16, 16, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ════════════════ ЛИСТ 3: СВОДКА ПО КАТЕГОРИЯМ ════════════════
    ws3 = wb.create_sheet("Сводка по категориям")
    h3 = ["Категория", "Выручка", "Транзакции", "Средний чек"]
    for c, h in enumerate(h3, 1):
        ws3.cell(row=1, column=c, value=h)
    style_header(ws3, 1, 4)
    write_data_rows(ws3, cat_summary, 2, {2: '#,##0', 3: '#,##0', 4: '#,##0.00'})
    lr3 = len(cat_summary) + 1
    ws3.auto_filter.ref = f"A1:D{lr3}"
    ws3.freeze_panes = "A2"
    ws3.conditional_formatting.add(
        f"A2:A{lr3}",
        FormulaRule(formula=[f"B2=MAX($B$2:$B${lr3})"],
                     fill=ORANGE_FILL, font=Font(bold=True, color="FFFFFF")),
    )
    for i, w in enumerate([16, 16, 16, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ════════════════ ЛИСТ 4: ДИНАМИКА ПО МЕСЯЦАМ ════════════════
    ws4 = wb.create_sheet("Динамика по месяцам")
    h4 = ["Месяц", "Выручка", "Транзакции"]
    for c, h in enumerate(h4, 1):
        ws4.cell(row=1, column=c, value=h)
    style_header(ws4, 1, 3)
    write_data_rows(ws4, monthly_list, 2, {2: '#,##0', 3: '#,##0'})
    for i, w in enumerate([10, 16, 16], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # ════════════════ ЛИСТ 5: ДАШБОРД ════════════════
    ws5 = wb.create_sheet("Дашборд")
    _build_dashboard(ws5, records, reg_summary, cat_summary, monthly_list,
                     total_revenue, total_transactions, len(records))

    return wb


def _build_dashboard(ws, records, reg_summary, cat_summary, monthly_list,
                     total_revenue, total_transactions, n_records):
    """Наполняет лист Дашборд."""
    for i, w in enumerate([22, 18, 18, 18, 4, 20, 18, 18, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ─── ЗАГОЛОВОК ───
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "ДАШБОРД ПРОДАЖ"
    c.font = TITLE_FONT
    c.fill = BLUE_FILL
    c.alignment = CENTER
    ws.row_dimensions[1].height = 40

    # ─── KPI ───
    avg_check = round(total_revenue / total_transactions, 2) if total_transactions else 0
    kpi_data = [
        ("B3", "Общая выручка, руб.", total_revenue, '#,##0'),
        ("D3", "Всего транзакций",   total_transactions, '#,##0'),
        ("F3", "Средний чек, руб.",  avg_check, '#,##0.00'),
        ("H3", "Регионов / Категорий", f"{len(REGIONS)} / {len(CATEGORIES)}", None),
    ]
    for cell_ref, label, value, nf in kpi_data:
        r = int(cell_ref[1])
        col_idx = ord(cell_ref[0]) - 64
        label_cell = ws.cell(row=r, column=col_idx, value=label)
        label_cell.font = KPI_LABEL_FONT
        label_cell.fill = LIGHT_BLUE_FILL
        label_cell.alignment = CENTER
        label_cell.border = THIN_BORDER

        val_cell = ws.cell(row=r + 1, column=col_idx, value=value)
        val_cell.font = KPI_VALUE_FONT
        val_cell.alignment = CENTER
        val_cell.border = THIN_BORDER
        if nf:
            val_cell.number_format = nf

    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 35

    # ─── ФИЛЬТРЫ ───
    ws["A6"] = "Фильтр по категории:"
    ws["A6"].font = BODY_BOLD
    ws["A6"].alignment = LEFT_CENTER

    dv_cat = DataValidation(
        type="list",
        formula1='"Все,' + ",".join(CATEGORIES) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_cat)
    dv_cat.add("B6")
    ws["B6"] = "Все"
    ws["B6"].font = BODY_FONT
    ws["B6"].alignment = CENTER
    ws["B6"].border = THIN_BORDER

    ws["D6"] = "Фильтр по региону:"
    ws["D6"].font = BODY_BOLD
    ws["D6"].alignment = LEFT_CENTER

    dv_reg = DataValidation(
        type="list",
        formula1='"Все,' + ",".join(REGIONS) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_reg)
    dv_reg.add("E6")
    ws["E6"] = "Все"
    ws["E6"].font = BODY_FONT
    ws["E6"].alignment = CENTER
    ws["E6"].border = THIN_BORDER

    # ─── ТАБЛИЦА: ВЫРУЧКА ПО РЕГИОНАМ ───
    ws["A8"] = "Выручка по регионам"
    ws["A8"].font = SECTION_FONT
    ws.merge_cells("A8:D8")

    for i, h in enumerate(["Регион", "Выручка", "Транзакции", "Средний чек"], 1):
        cell = ws.cell(row=9, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = BLUE_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    reg_end = 9 + len(reg_summary)
    for idx, (reg, rev, trx, avg) in enumerate(reg_summary):
        r = 10 + idx
        ws.cell(row=r, column=1, value=reg).font = BODY_BOLD
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=1).border = THIN_BORDER

        formula_rev = (
            f'=IF($B$6="Все",SUMIF(Данные!B:B,A{r},Данные!D:D),'
            f'SUMIFS(Данные!D:D,Данные!B:B,A{r},Данные!C:C,$B$6))'
        )
        formula_trx = (
            f'=IF($B$6="Все",SUMIF(Данные!B:B,A{r},Данные!E:E),'
            f'SUMIFS(Данные!E:E,Данные!B:B,A{r},Данные!C:C,$B$6))'
        )
        for c, val, nf in [(2, formula_rev, '#,##0'),
                           (3, formula_trx, '#,##0'),
                           (4, '=IFERROR(B{r}/C{r},0)', '#,##0.00')]:
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            cell.number_format = nf

    ws.conditional_formatting.add(
        f"B10:B{reg_end}",
        FormulaRule(formula=[f"B10=MAX($B$10:$B${reg_end})"],
                     fill=ORANGE_FILL, font=Font(bold=True, color="FFFFFF")),
    )

    # ─── ТАБЛИЦА: ВЫРУЧКА ПО КАТЕГОРИЯМ ───
    ws["F8"] = "Выручка по категориям"
    ws["F8"].font = SECTION_FONT
    ws.merge_cells("F8:I8")

    for i, h in enumerate(["Категория", "Выручка", "Транзакции", "Средний чек"], 6):
        cell = ws.cell(row=9, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = BLUE_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    cat_end = 9 + len(cat_summary)
    for idx, (cat, rev, trx, avg) in enumerate(cat_summary):
        r = 10 + idx
        ws.cell(row=r, column=6, value=cat).font = BODY_BOLD
        ws.cell(row=r, column=6).alignment = CENTER
        ws.cell(row=r, column=6).border = THIN_BORDER

        formula_rev = (
            f'=IF($E$6="Все",SUMIF(Данные!C:C,F{r},Данные!D:D),'
            f'SUMIFS(Данные!D:D,Данные!C:C,F{r},Данные!B:B,$E$6))'
        )
        formula_trx = (
            f'=IF($E$6="Все",SUMIF(Данные!C:C,F{r},Данные!E:E),'
            f'SUMIFS(Данные!E:E,Данные!C:C,F{r},Данные!B:B,$E$6))'
        )
        for c, val, nf in [(7, formula_rev, '#,##0'),
                           (8, formula_trx, '#,##0'),
                           (9, '=IFERROR(G{r}/H{r},0)', '#,##0.00')]:
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            cell.number_format = nf

    ws.conditional_formatting.add(
        f"G10:G{cat_end}",
        FormulaRule(formula=[f"G10=MAX($G$10:$G${cat_end})"],
                     fill=ORANGE_FILL, font=Font(bold=True, color="FFFFFF")),
    )

    # ─── МЕСЯЧНЫЕ ДАННЫЕ ДЛЯ ГРАФИКА ───
    ws["A16"] = "Динамика по месяцам"
    ws["A16"].font = SECTION_FONT
    ws.merge_cells("A16:C16")

    for i, h in enumerate(["Месяц", "Выручка", "Транзакции"], 1):
        cell = ws.cell(row=17, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = BLUE_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for idx, (m, rev, trx) in enumerate(monthly_list):
        r = 18 + idx
        ws.cell(row=r, column=1, value=m).font = BODY_FONT
        ws.cell(row=r, column=1).alignment = CENTER
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=rev).font = BODY_FONT
        ws.cell(row=r, column=2).alignment = CENTER
        ws.cell(row=r, column=2).number_format = '#,##0'
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=3, value=trx).font = BODY_FONT
        ws.cell(row=r, column=3).alignment = CENTER
        ws.cell(row=r, column=3).number_format = '#,##0'
        ws.cell(row=r, column=3).border = THIN_BORDER

    mon_end = 17 + len(monthly_list)

    # ─── ДИАГРАММЫ ───
    # Линейная: динамика выручки
    chart1 = LineChart()
    chart1.title = "Динамика выручки по месяцам"
    chart1.style = 10
    chart1.y_axis.title = "Руб."
    chart1.width = 24
    chart1.height = 14

    d1 = Reference(ws, min_col=2, min_row=17, max_row=mon_end, max_col=2)
    c1 = Reference(ws, min_col=1, min_row=18, max_row=mon_end)
    chart1.add_data(d1, titles_from_data=True)
    chart1.set_categories(c1)
    chart1.series[0].graphicalProperties.line.solidFill = BLUE
    chart1.series[0].graphicalProperties.line.width = 28000

    d2 = Reference(ws, min_col=3, min_row=17, max_row=mon_end, max_col=3)
    chart1.add_data(d2, titles_from_data=True)
    chart1.series[1].graphicalProperties.line.solidFill = ORANGE
    chart1.series[1].graphicalProperties.line.width = 28000
    chart1.legend = None
    ws.add_chart(chart1, "A31")

    # Столбчатая: выручка по регионам
    chart2 = BarChart()
    chart2.title = "Выручка по регионам"
    chart2.style = 10
    chart2.y_axis.title = "Руб."
    chart2.width = 20
    chart2.height = 14

    d3 = Reference(ws, min_col=2, min_row=9, max_row=reg_end, max_col=2)
    c3 = Reference(ws, min_col=1, min_row=10, max_row=reg_end)
    chart2.add_data(d3, titles_from_data=True)
    chart2.set_categories(c3)
    chart2.series[0].graphicalProperties.solidFill = BLUE
    chart2.legend = None
    ws.add_chart(chart2, "F31")

    # Круговая: доля категорий
    chart3 = PieChart()
    chart3.title = "Доля категорий в выручке"
    chart3.width = 20
    chart3.height = 14

    d4 = Reference(ws, min_col=7, min_row=9, max_row=cat_end, max_col=7)
    c4 = Reference(ws, min_col=6, min_row=10, max_row=cat_end)
    chart3.add_data(d4, titles_from_data=True)
    chart3.set_categories(c4)

    pie_colors = [BLUE, ORANGE, "4CAF50", "E91E63", "9C27B0"]
    for i, clr in enumerate(pie_colors):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = clr
        chart3.series[0].data_points.append(pt)

    chart3.dataLabels = DataLabelList()
    chart3.dataLabels.showPercent = True
    chart3.dataLabels.showCatName = True
    ws.add_chart(chart3, "A49")


# ────────────────────────── MAIN ──────────────────────────
def main():
    print("Генерация данных о продажах...")
    sys.stdout.flush()
    records = generate_data()
    print(f"Сгенерировано {len(records)} записей")
    sys.stdout.flush()

    print("Расчёт сводок...")
    sys.stdout.flush()
    reg_summary, cat_summary, monthly_list, total_rev, total_trx = \
        compute_summaries(records)
    print(f"Общая выручка: {total_rev:,.0f} руб.")
    print(f"Всего транзакций: {total_trx:,}")
    print(f"Средний чек: {total_rev / total_trx:,.2f} руб.")
    sys.stdout.flush()

    print("Создание Excel-файла...")
    sys.stdout.flush()
    wb = create_workbook(records, reg_summary, cat_summary, monthly_list,
                         total_rev, total_trx)

    out = "sales_dashboard.xlsx"
    try:
        wb.save(out)
        print(f"Файл сохранён: {os.path.abspath(out)}")
    except PermissionError:
        print(f"ОШИБКА: Файл '{out}' занят. Закройте его и повторите.")
        sys.exit(1)
    except Exception as e:
        print(f"ОШИБКА при сохранении: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
